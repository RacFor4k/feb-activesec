import csv
import os
from openpyxl import Workbook
from statistics import pstdev

ROOT_DIR = '.\\statistics\\'

def main():
    wb = Workbook()
    wb.remove(wb.active)

    meta = [['Type', 'entropy mean', 'entropy median', 'entropy stdev', 'chi2 mean', 'chi2 median', 'chi2 stdev']]
    
    for file in [i for i in os.listdir(ROOT_DIR) if os.path.isfile(os.path.join(ROOT_DIR, i))]:
        data = []
        ws = wb.create_sheet('-'.join(file.split('-')[:-1]))
        with open(os.path.join(ROOT_DIR,file), newline='') as f:
            reader = csv.reader(f)
            rows = [[row[0]]+[float(cell) for cell in row[1:]] for row in reader] 
                
            columns = list(zip(*rows))
            
            for row in rows:
                data.append(row)
            
            size_mean = sum(columns[1])/len(rows)
            size_median = sorted(columns[1])[len(rows)//2]
            size_stdev = pstdev(columns[1])
            
            entropy_mean = sum(columns[2])/len(rows)
            entropy_median = sorted(columns[2])[len(rows)//2]
            entropy_stdev = pstdev(columns[2])
            
            chi2_mean = sum(columns[3])/len(rows)
            chi2_median = sorted(columns[3])[len(rows)//2]
            chi2_stdev = pstdev(columns[3])
            
            data.append(['mean', f'{size_mean/1024:.2f}', entropy_mean, chi2_mean])
            data.append(['median', f'{size_median/1024:.2f}', entropy_median, chi2_median])
            data.append(['stdev', f'{size_stdev/1024:.2f}', entropy_stdev, chi2_stdev])
            
            
            meta.append([file, entropy_mean, entropy_median, entropy_stdev, chi2_mean, chi2_median, chi2_stdev])
        for i in data:
            ws.append(i)
    ws = wb.create_sheet("Meta")
    for i in meta:
        ws.append(i)
    wb.save('processed.xlsx')

if __name__ == '__main__':
    main()